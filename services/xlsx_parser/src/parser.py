import pandas as pd
import os
import re
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from typing import Optional, List, Dict

from models import (
    IncomingDirection, ExcelDataFile, Student,
    FormatControl, FormatRetests, StudyProgram, ControlTable,
    Specialization, Base
)


# ══════════════════════════════════════════════
#  УТИЛИТЫ
# ══════════════════════════════════════════════

def get_or_create(session: Session, model, defaults: dict = None, **kwargs):
    """Найти запись или создать новую"""
    instance = session.query(model).filter_by(**kwargs).first()
    if not instance:
        params = {**kwargs}
        if defaults:
            params.update(defaults)
        instance = model(**params)
        session.add(instance)
        session.flush()
    return instance


def parse_code_from_filename(filename: str) -> str:
    """
    '0001_Абдулаев_Магомедали_..._OZ.xlsx' → '0001'
    """
    basename = os.path.splitext(filename)[0]
    match = re.match(r'^(\d+)', basename)
    if match:
        return match.group(1)
    raise ValueError(f"Не удалось извлечь code_file из имени файла: {filename}")


# ══════════════════════════════════════════════
#  ПАРСИНГ ЛИСТА "ТИТУЛ"
# ══════════════════════════════════════════════

def parse_title_sheet(filepath: str) -> dict:
    result = {
        'direction_name': None,
        'full_name': None,
    }

    wb = load_workbook(filepath, data_only=True)

    title_sheet_name = None
    for name in wb.sheetnames:
        if 'титул' in name.lower():
            title_sheet_name = name
            break

    if not title_sheet_name:
        title_sheet_name = wb.sheetnames[0]
        print(f"   ⚠️ Лист 'Титул' не найден, используем '{title_sheet_name}'")

    ws = wb[title_sheet_name]

    all_values = []
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val and isinstance(val, str):
                all_values.append((cell.row, cell.column, val.strip()))

    # ── Паттерны для ФИО ──
    fio_patterns = [
        # "индивидуальный план Абдулаев Магомедали Абдурахманович"
        # (с отчеством)
        r'[Ии]ндивидуальн\w+\s+план\s+'
        r'([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)',

        # "индивидуальный план Абдулаев Магомедали"
        # (без отчества)
        r'[Ии]ндивидуальн\w+\s+план\s+'
        r'([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)',

        # "Студент: Фамилия Имя Отчество"
        r'(?:Студент|Обучающ\w+|ФИО)\s*:?\s*'
        r'([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)',

        # "Студент: Фамилия Имя"
        r'(?:Студент|Обучающ\w+|ФИО)\s*:?\s*'
        r'([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)',
    ]

    for row_num, col_num, val_clean in all_values:
        # ── Направление ──
        if 'Направление подготовки' in val_clean and not result['direction_name']:
            result['direction_name'] = (
                re.sub(
                    r'\s+', ' ',
                    val_clean
                    .replace('\n', ' ')
                    .replace('_x000D_', '')
                    .replace('\r', ' ')
                ).strip()
            )

        # ── ФИО ──
        if not result['full_name']:
            for pattern in fio_patterns:
                m = re.search(pattern, val_clean)
                if m:
                    candidate = m.group(1).strip()
                    if len(candidate) > 3:
                        result['full_name'] = candidate
                        break

    wb.close()
    return result


# ══════════════════════════════════════════════
#  ПАРСИНГ ЛИСТА "ПЕРЕАТТЕСТАЦИЯ"
# ══════════════════════════════════════════════

def parse_reattest_sheet(filepath: str) -> List[Dict]:
    """Парсит таблицу с листа Переаттестация"""
    wb_temp = load_workbook(filepath, read_only=True)
    sheet_name = None
    for name in wb_temp.sheetnames:
        if 'переаттестац' in name.lower():
            sheet_name = name
            break
    wb_temp.close()

    if not sheet_name:
        raise ValueError(f"Лист 'Переаттестация' не найден")

    df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')

    top_headers = df.columns.tolist()
    sub_headers = df.iloc[1].tolist()

    column_structure = []
    current_group = None

    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        top_str = str(top)
        sub_str = str(sub) if pd.notna(sub) else None

        if top_str.startswith('Unnamed'):
            group = current_group
        elif top_str == '-' or top_str.startswith('-.'):
            group = None
            current_group = None
        else:
            group = top_str
            current_group = top_str

        if group and sub_str:
            flat_key = f"{group}__{sub_str}"
        elif sub_str:
            flat_key = sub_str
        else:
            flat_key = f"col_{i}"

        column_structure.append({
            "index": i,
            "group": group,
            "name": sub_str,
            "flat_key": flat_key,
        })

    data_df = df.iloc[2:].copy().reset_index(drop=True)

    mask_itogo = data_df.iloc[:, 1].astype(str).str.contains('Итого', na=False)
    data_df = data_df[~mask_itogo]
    data_df = data_df.dropna(how='all')

    data_df.iloc[:, 0] = data_df.iloc[:, 0].ffill()
    data_df.iloc[:, 1] = data_df.iloc[:, 1].ffill()

    data_df = data_df[data_df.iloc[:, 1].notna()]
    data_df = data_df.reset_index(drop=True)

    rows = []
    for _, row in data_df.iterrows():
        row_dict = {}
        for col_info in column_structure:
            val = row.iloc[col_info["index"]]
            if pd.isna(val):
                val = None
            elif hasattr(val, 'item'):
                val = val.item()
            row_dict[col_info["flat_key"]] = val
        rows.append(row_dict)

    return rows


# ══════════════════════════════════════════════
#  СОХРАНЕНИЕ В БД
# ══════════════════════════════════════════════

def save_excel_to_db(filepath: str, session: Session):
    """
    Логика:
    1. code_file — из имени файла
    2. ФИО — ТОЛЬКО из листа Титул
    3. Студент ищется по ФИО:
       - найден + code_file совпадает → привязываем direction
       - найден + code_file другой → обновляем
       - не найден → создаём
    4. Данные переаттестации → ControlTable
    """
    filename = os.path.basename(filepath)

    # ═══ 1. Код из имени файла ═══
    code_file = parse_code_from_filename(filename)
    print(f"📄 {filename}")
    print(f"   Code: {code_file}")

    # ═══ 2. Проверяем дубликат ═══
    existing_file = session.query(ExcelDataFile).filter_by(
        code_file=int(code_file)
    ).first()
    if existing_file:
        print(f"   ⚠️ Файл уже загружен (code_file={code_file}), пропускаем\n")
        return existing_file

    # ═══ 3. Парсим Титул (ФИО только отсюда!) ═══
    title_data = parse_title_sheet(filepath)

    direction_name = title_data['direction_name']
    full_name = title_data['full_name']

    if not direction_name:
        raise ValueError(
            f"Направление не найдено на листе Титул в файле {filename}"
        )

    if not full_name:
        raise ValueError(
            f"ФИО не найдено на листе Титул в файле {filename}. "
            f"Запустите debug_sheet('{filepath}', 'Титул') чтобы "
            f"увидеть содержимое и поправить паттерн поиска."
        )

    print(f"   Направление: {direction_name[:80]}...")
    print(f"   ФИО (из Титула): {full_name}")

    # ═══ 4. IncomingDirection ═══
    direction = get_or_create(
        session, IncomingDirection,
        name=direction_name
    )
    print(f"   Direction ID: {direction.id}")

    # ═══ 5. Student ═══
    student = session.query(Student).filter_by(
        full_name=full_name
    ).first()

    if student:
        if student.file_code == code_file:
            # Тот же студент, тот же файл — обновляем direction
            print(f"   Студент найден (id={student.id}), code_file совпадает")
            if student.incoming_direction_id != direction.id:
                student.incoming_direction_id = direction.id
                print(f"   → Обновлён incoming_direction_id → {direction.id}")
        else:
            # Такое же ФИО но другой code_file — это ДРУГОЙ файл
            # для того же студента или полный тёзка
            print(
                f"   Студент '{full_name}' найден (id={student.id}), "
                f"но code_file отличается: БД='{student.file_code}', "
                f"файл='{code_file}'"
            )
            # Обновляем
            student.file_code = code_file
            student.file_name = filename
            student.incoming_direction_id = direction.id
            print(f"   → Обновлён code_file и direction")
    else:
        # Создаём нового студента
        # default_spec = get_or_create(
        #     session, Specialization,
        #     name="Не определена"
        # )
        student = Student(
            full_name=full_name,
            file_code=code_file,
            file_name=filename,
            specialization_id=None,
            incoming_direction_id=direction.id,
        )
        session.add(student)
        session.flush()
        print(f"   Студент создан (id={student.id})")

    
    
    # ═══ 6. ExcelDataFile ═══
    excel_file = session.query(ExcelDataFile).filter_by(
        code_file=code_file
    ).first()
    if not excel_file:
        excel_file = ExcelDataFile(
            name=filename,
            full_name=full_name,
            code_file=code_file,
            incoming_direction_id=direction.id,
        )
        session.add(excel_file)
        session.flush()
        print(f"   ExcelDataFile создан (id={excel_file.id})")

    # ═══ 7. Парсим Переаттестация ═══
    rows_data = parse_reattest_sheet(filepath)
    print(f"   Строк в таблице: {len(rows_data)}")

    # ═══ 8. ControlTable ═══
    def safe_str(val):
        if val is None:
            return None
        try:
            return str(int(float(val)))
        except (ValueError, TypeError):
            return str(val)

    created_count = 0
    for row_data in rows_data:
        naimenovanie = row_data.get('Наименование')
        if not naimenovanie:
            continue

        program = get_or_create(
            session, StudyProgram,
            name=str(naimenovanie).strip()
        )

        format_retests = None
        fr_val = row_data.get('Зачет результатов обучения')
        if fr_val:
            format_retests = get_or_create(
                session, FormatRetests,
                format_name=str(fr_val).strip()
            )

        fc_norma = None
        fc_norma_val = row_data.get('Форма пром. атт.')
        if fc_norma_val:
            fc_norma = get_or_create(
                session, FormatControl,
                format_name=str(fc_norma_val).strip()
            )

        hours_normal = row_data.get('По плану__Часов')
        hours_fact = row_data.get('Изучено и зачтено__Часов')

        control = ControlTable(
            incoming_direction_id=direction.id,
            study_program_id=program.id,
            format_control_norma_id=fc_norma.id if fc_norma else None,
            format_control_fact_id=None,
            format_retests_id=format_retests.id if format_retests else None,
            hours_normal=safe_str(hours_normal),
            hours_fact=safe_str(hours_fact),
        )
        session.add(control)
        created_count += 1

    session.commit()
    print(f"   ✅ Записано {created_count} строк в ControlTable\n")
    return excel_file


# ══════════════════════════════════════════════
#  ОБРАБОТКА ПАПКИ
# ══════════════════════════════════════════════

def process_all_files(directory: str, session: Session):
    files = sorted([
        f for f in os.listdir(directory)
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')
    ])

    print(f"{'=' * 60}")
    print(f"Найдено файлов: {len(files)}")
    print(f"{'=' * 60}\n")

    success = 0
    errors = []

    for filename in files:
        filepath = os.path.join(directory, filename)
        try:
            save_excel_to_db(filepath, session)
            success += 1
        except Exception as e:
            print(f"   ❌ ОШИБКА: {e}\n")
            errors.append((filename, str(e)))
            session.rollback()

    print(f"\n{'=' * 60}")
    print(f"Успешно: {success}/{len(files)}")
    if errors:
        print(f"Ошибки:")
        for fname, err in errors:
            print(f"  - {fname}: {err}")
    print(f"{'=' * 60}")


# ══════════════════════════════════════════════
#  ОТЛАДКА
# ══════════════════════════════════════════════

def debug_sheet(filepath: str, sheet_name: str = 'Титул'):
    """Вывести все непустые ячейки — помогает найти где ФИО"""
    wb = load_workbook(filepath, data_only=True)

    target = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower():
            target = name
            break

    if not target:
        print(f"Лист '{sheet_name}' не найден. Доступные: {wb.sheetnames}")
        return

    ws = wb[target]
    print(f"\n{'=' * 70}")
    print(f"ЛИСТ: {target}  |  ФАЙЛ: {os.path.basename(filepath)}")
    print(f"{'=' * 70}")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val_preview = str(cell.value).replace('\n', '\\n')[:100]
                print(
                    f"  [{cell.row:3d}, {cell.column:2d}] "
                    f"{cell.coordinate:6s}: {val_preview}"
                )

    wb.close()


# ══════════════════════════════════════════════
#  ЗАПУСК
# ══════════════════════════════════════════════

if __name__ == '__main__':
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    engine = create_engine('sqlite:///data.db')
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine)

    # debug_sheet('./inp/0001_Абдулаев_Магомедали_Абдурахманович_Uch_plan_38_03_01_FiK_OZ.xlsx', 'Титул')

    with SessionLocal() as session:
        process_all_files('./inp/', session)