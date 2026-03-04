import pandas as pd
import os
import re
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from typing import Optional, List, Dict

from models import (
    IncomingDirection, ExcelDataFile, Student,
    FormatControl, FormatRetests, StudyProgram, ControlTable
)


# ══════════════════════════════════════════════
#  УТИЛИТЫ
# ══════════════════════════════════════════════

def get_or_create(session: Session, model, **kwargs):
    """Найти запись или создать новую"""
    instance = session.query(model).filter_by(**kwargs).first()
    if not instance:
        instance = model(**kwargs)
        session.add(instance)
        session.flush()  # чтобы получить id
    return instance


# ══════════════════════════════════════════════
#  ПАРСИНГ ЛИСТА "ТИТУЛ"
# ══════════════════════════════════════════════

def parse_title_sheet(filepath: str) -> dict:
    """
    Извлекает с листа Титул:
    - direction_name: направление подготовки
    - full_name: ФИО студента
    """
    result = {
        'direction_name': None,
        'full_name': None,
    }

    # ── Способ 1: через openpyxl (надёжнее для merged cells) ──
    wb = load_workbook(filepath, data_only=True)

    # Определяем имя листа (может быть разное написание)
    title_sheet_name = None
    for name in wb.sheetnames:
        if 'титул' in name.lower():
            title_sheet_name = name
            break

    if not title_sheet_name:
        raise ValueError(f"Лист 'Титул' не найден. Листы: {wb.sheetnames}")

    ws = wb[title_sheet_name]

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not val or not isinstance(val, str):
                continue

            val_clean = val.strip()

            # ── Направление подготовки ──
            if 'Направление подготовки' in val_clean:
                result['direction_name'] = (
                    re.sub(r'\s+', ' ',
                           val_clean
                           .replace('\n', ' ')
                           .replace('_x000D_', '')
                           .replace('\r', ' '))
                    .strip()
                )

            # ── ФИО студента ──
            # Ищем по ключевым словам
            fio_patterns = [
                r'(?:Студент|Обучающ\w+|ФИО)\s*:?\s*'
                r'([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)',
                r'(?:Студент|Обучающ\w+)\s*:?\s*(.+)',
            ]
            for pattern in fio_patterns:
                m = re.search(pattern, val_clean)
                if m and not result['full_name']:
                    candidate = m.group(1).strip()
                    if len(candidate) > 5:  # не инициалы
                        result['full_name'] = candidate

    wb.close()

    # ── Способ 2 (fallback): через pandas ──
    if not result['direction_name'] or not result['full_name']:
        df = pd.read_excel(
            filepath,
            sheet_name=title_sheet_name,
            header=None,
            engine='openpyxl'
        )
        for _, row in df.iterrows():
            for val in row:
                if pd.isna(val):
                    continue
                val_str = str(val).strip()

                if not result['direction_name'] and 'Направление подготовки' in val_str:
                    result['direction_name'] = re.sub(
                        r'\s+', ' ',
                        val_str.replace('\n', ' ')
                    ).strip()

                if not result['full_name']:
                    for pat in fio_patterns:
                        m = re.search(pat, val_str)
                        if m:
                            result['full_name'] = m.group(1).strip()
                            break

    return result


# ══════════════════════════════════════════════
#  ПАРСИНГ ЛИСТА "ПЕРЕАТТЕСТАЦИЯ"
# ══════════════════════════════════════════════

def parse_reattest_sheet(filepath: str) -> List[Dict]:
    """
    Парсит таблицу с листа Переаттестация.
    Возвращает список словарей — каждый словарь = одна строка таблицы.
    """
    # ── Ищем лист ──
    wb_temp = load_workbook(filepath, read_only=True)
    sheet_name = None
    for name in wb_temp.sheetnames:
        if 'переаттестац' in name.lower():
            sheet_name = name
            break
    wb_temp.close()

    if not sheet_name:
        raise ValueError(f"Лист 'Переаттестация' не найден")

    df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')

    # ── Строим маппинг колонок ──
    top_headers = df.columns.tolist()
    sub_headers = df.iloc[1].tolist()  # строка с реальными заголовками

    column_structure = []
    current_group = None

    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        top_str = str(top)
        sub_str = str(sub) if pd.notna(sub) else None

        # Определяем группу
        if top_str.startswith('Unnamed'):
            # Часть объединённой ячейки → та же группа
            group = current_group
        elif top_str == '-' or top_str.startswith('-.'):
            # Одиночная колонка
            group = None
            current_group = None
        else:
            # Новая группа (По плану, Изучено и зачтено, ...)
            group = top_str
            current_group = top_str

        # Формируем плоский ключ
        if group and sub_str:
            flat_key = f"{group}__{sub_str}"
        elif sub_str:
            flat_key = sub_str
        else:
            flat_key = f"col_{i}"

        column_structure.append({
            "index": i,
            "group": group,
            "name": sub_str,
            "flat_key": flat_key,
        })

    # ── Извлекаем данные (пропускаем строки заголовков) ──
    data_df = df.iloc[2:].copy().reset_index(drop=True)

    # Убираем строки "Итого"
    mask_itogo = data_df.iloc[:, 1].astype(str).str.contains('Итого', na=False)
    data_df = data_df[~mask_itogo]

    # Убираем полностью пустые строки
    data_df = data_df.dropna(how='all')

    # Forward-fill Индекс и Наименование
    # (для предметов на несколько семестров, например Б1.В.01)
    data_df.iloc[:, 0] = data_df.iloc[:, 0].ffill()  # Индекс
    data_df.iloc[:, 1] = data_df.iloc[:, 1].ffill()  # Наименование

    # Убираем строки где и после ffill нет Наименования
    data_df = data_df[data_df.iloc[:, 1].notna()]
    data_df = data_df.reset_index(drop=True)

    # ── Формируем список словарей ──
    rows = []
    for _, row in data_df.iterrows():
        row_dict = {}
        for col_info in column_structure:
            val = row.iloc[col_info["index"]]
            if pd.isna(val):
                val = None
            elif hasattr(val, 'item'):
                val = val.item()
            row_dict[col_info["flat_key"]] = val
        rows.append(row_dict)

    return rows


# ══════════════════════════════════════════════
#  СОХРАНЕНИЕ В БД
# ══════════════════════════════════════════════

def save_excel_to_db(
    filepath: str,
    session: Session,
    code_file: int = None,
    manual_fio: str = None,
):
    """
    Обрабатывает один Excel файл и сохраняет в БД.

    Args:
        filepath: путь к .xlsx файлу
        session: SQLAlchemy сессия
        code_file: уникальный код файла (если None — генерируется из имени)
        manual_fio: ФИО вручную (если автоопределение не работает)
    """
    filename = os.path.basename(filepath)

    # ═══ 1. Парсим Титул ═══
    title_data = parse_title_sheet(filepath)

    direction_name = title_data['direction_name']
    full_name = manual_fio or title_data['full_name'] or filename

    if not direction_name:
        raise ValueError(f"Направление не найдено в файле {filename}")

    print(f"📄 {filename}")
    print(f"   Направление: {direction_name[:80]}...")
    print(f"   ФИО: {full_name}")

    # ═══ 2. IncomingDirection (get or create) ═══
    direction = get_or_create(
        session, IncomingDirection,
        name=direction_name
    )

    # ═══ 3. ExcelDataFile ═══
    if code_file is None:
        code_file = abs(hash(filepath + filename)) % (10 ** 9)

    existing_file = session.query(ExcelDataFile).filter_by(
        code_file=code_file
    ).first()

    if existing_file:
        print(f"   ⚠️ Файл уже загружен (code_file={code_file}), пропускаем")
        return existing_file

    excel_file = ExcelDataFile(
        name=filename,
        full_name=full_name,
        code_file=code_file,
        incoming_direction_id=direction.id,
    )
    session.add(excel_file)
    session.flush()

    # ═══ 4. Парсим Переаттестация ═══
    rows_data = parse_reattest_sheet(filepath)
    print(f"   Строк в таблице: {len(rows_data)}")

    # ═══ 5. Сохраняем каждую строку ═══
    for row_data in rows_data:
        naimenovanie = row_data.get('Наименование')
        if not naimenovanie:
            continue

        # StudyProgram
        program = get_or_create(
            session, StudyProgram,
            name=str(naimenovanie).strip()
        )

        # FormatRetests (Переаттестовано(частично) и т.д.)
        format_retests = None
        fr_val = row_data.get('Зачет результатов обучения')
        if fr_val:
            format_retests = get_or_create(
                session, FormatRetests,
                format_name=str(fr_val).strip()
            )

        # FormatControl NORMA (Экзамен / Зачет / Зачет с оценкой)
        fc_norma = None
        fc_val = row_data.get('Форма пром. атт.')
        if fc_val:
            fc_norma = get_or_create(
                session, FormatControl,
                format_name=str(fc_val).strip()
            )

        # Часы
        hours_normal = row_data.get('По плану__Часов')
        hours_fact = row_data.get('Изучено и зачтено__Часов')

        control = ControlTable(
            incoming_direction_id=direction.id,
            study_program_id=program.id,
            format_control_norma_id=fc_norma.id if fc_norma else None,
            format_control_fact_id=None,  # заполнить при необходимости
            format_retests_id=format_retests.id if format_retests else None,
            hours_normal=str(int(hours_normal)) if hours_normal else None,
            hours_fact=str(int(hours_fact)) if hours_fact else None,
        )
        session.add(control)

    session.commit()
    print(f"   ✅ Сохранено успешно\n")

    return excel_file


# ══════════════════════════════════════════════
#  ОБРАБОТКА ВСЕХ ФАЙЛОВ В ПАПКЕ
# ══════════════════════════════════════════════

def process_all_files(directory: str, session: Session):
    """Обработать все .xlsx файлы в папке"""
    files = sorted([
        f for f in os.listdir(directory)
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')
    ])

    print(f"{'=' * 60}")
    print(f"Найдено файлов: {len(files)}")
    print(f"{'=' * 60}\n")

    success = 0
    errors = []

    for filename in files:
        filepath = os.path.join(directory, filename)
        try:
            save_excel_to_db(filepath, session)
            success += 1
        except Exception as e:
            print(f"   ❌ ОШИБКА: {e}\n")
            errors.append((filename, str(e)))
            session.rollback()

    # ── Итоги ──
    print(f"\n{'=' * 60}")
    print(f"Успешно: {success}/{len(files)}")
    if errors:
        print(f"Ошибки:")
        for fname, err in errors:
            print(f"  - {fname}: {err}")
    print(f"{'=' * 60}")


# ══════════════════════════════════════════════
#  ОТЛАДКА: посмотреть содержимое листа
# ══════════════════════════════════════════════

def debug_sheet(filepath: str, sheet_name: str = 'Титул'):
    """
    Вывести все непустые ячейки листа.
    Помогает найти где лежит ФИО и направление.
    """
    wb = load_workbook(filepath, data_only=True)

    # Найти лист
    target = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower():
            target = name
            break

    if not target:
        print(f"Лист '{sheet_name}' не найден. Доступные: {wb.sheetnames}")
        return

    ws = wb[target]
    print(f"\n{'=' * 70}")
    print(f"ЛИСТ: {target}  |  ФАЙЛ: {os.path.basename(filepath)}")
    print(f"{'=' * 70}")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                val_preview = str(cell.value).replace('\n', '\\n')[:100]
                print(f"  [{cell.row:3d}, {cell.column:2d}] "
                      f"{cell.coordinate:6s}: {val_preview}")

    wb.close()


# ══════════════════════════════════════════════
#  ЗАПУСК
# ══════════════════════════════════════════════

if __name__ == '__main__':
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from models import Base
    
    engine = create_engine('sqlite:///dbname.db')
    SessionLocal = sessionmaker(bind=engine)
    Base.metadata.create_all(engine)
    # ── Отладка: посмотреть что в файле ──
    # debug_sheet('path/to/file.xlsx', 'Титул')
    # debug_sheet('path/to/file.xlsx', 'Переаттестация')

    # ── Один файл ──
    # with SessionLocal() as session:
    #     save_excel_to_db('path/to/file.xlsx', session)

    # ── Все файлы в папке ──
    with SessionLocal() as session:
        process_all_files('D:/Develop/python/ocr-diploma-sheduler/services/xlsx_processor/src/inp/', session)