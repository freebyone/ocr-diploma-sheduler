import re
from typing import Optional, Dict, List, Tuple, Any
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker
import logging

# Исправленные импорты для работы в структуре src
from src.models import (
    Base, Direction, Specialization, 
    StudyProgram, FormatControl, FormatRetests, ControlTable
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class XLSXParser:
    def __init__(self, file_path: str, db_url: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        
        # Настройка подключения к БД
        self.engine = create_engine(db_url)
        self.Session = sessionmaker(bind=self.engine)
        
        # Создаем таблицы, если их нет
        Base.metadata.create_all(self.engine)
    
    def _clean_text(self, text: str) -> str:
        """Очистка текста от HTML тегов и специальных символов"""
        if not text:
            return ""
        
        # Удаляем HTML теги
        text = re.sub(r'<[^>]+>', '', str(text))
        # Заменяем специальные символы
        text = text.replace('_x000d_', ' ').replace('\n', ' ').replace('\r', ' ')
        # Удаляем лишние пробелы
        text = re.sub(r'\s+', ' ', text).strip()
        
        return text
    
    def _get_or_create_direction(self, name: str) -> Optional[Direction]:
        """Получить или создать направление"""
        if not name:
            return None
            
        cleaned_name = self._clean_text(name)
        if not cleaned_name:
            return None
            
        # Создаем новую сессию
        session = self.Session()
        try:
            direction = session.execute(
                select(Direction).where(Direction.name == cleaned_name)
            ).scalar_one_or_none()
            
            if not direction:
                direction = Direction(name=cleaned_name)
                session.add(direction)
                session.commit()
                logger.info(f"Создано новое направление: {cleaned_name}")
            
            # Возвращаем копию данных, а не объект, привязанный к сессии
            result = {
                'id': direction.id,
                'name': direction.name
            }
            return result
        finally:
            session.close()
    
    def _get_or_create_specialization(
        self, 
        name: str, 
        code: str,
        direction_id: Optional[int] = None
    ) -> Optional[Dict]:
        """Получить или создать специализацию (без привязки к вузу)"""
        cleaned_name = self._clean_text(name)
        cleaned_code = self._clean_text(code) if code else ""
        
        if not cleaned_name:
            return None
            
        # Создаем новую сессию
        session = self.Session()
        try:
            specialization = session.execute(
                select(Specialization).where(
                    Specialization.name == cleaned_name,
                    Specialization.code == cleaned_code
                )
            ).scalar_one_or_none()
            
            if not specialization:
                specialization = Specialization(
                    name=cleaned_name,
                    code=cleaned_code,
                    direction_id=direction_id,
                    university_id=None
                )
                session.add(specialization)
                session.commit()
                logger.info(f"Создана новая специализация: {cleaned_name} ({cleaned_code})")
            else:
                # Обновляем direction_id, если нужно
                if direction_id and specialization.direction_id != direction_id:
                    specialization.direction_id = direction_id
                    session.commit()
                    logger.info(f"Обновлен direction_id для специализации {cleaned_name}")
            
            # Возвращаем копию данных
            result = {
                'id': specialization.id,
                'name': specialization.name,
                'code': specialization.code,
                'direction_id': specialization.direction_id
            }
            return result
        finally:
            session.close()
    
    def _get_or_create_study_program(self, name: str) -> Optional[Dict]:
        """Получить или создать учебную программу (предмет)"""
        if not name:
            return None
            
        cleaned_name = self._clean_text(name)
        if not cleaned_name or cleaned_name in ["Наименование", "Индекс", "-"]:
            return None
            
        session = self.Session()
        try:
            program = session.execute(
                select(StudyProgram).where(StudyProgram.name == cleaned_name)
            ).scalar_one_or_none()
            
            if not program:
                program = StudyProgram(name=cleaned_name)
                session.add(program)
                session.commit()
                logger.info(f"Создана новая учебная программа: {cleaned_name}")
            
            result = {
                'id': program.id,
                'name': program.name
            }
            return result
        finally:
            session.close()
    
    def _get_or_create_format_control(self, format_name: str) -> Optional[Dict]:
        """Получить или создать формат контроля"""
        if not format_name:
            return None
            
        cleaned_name = self._clean_text(format_name)
        if not cleaned_name or cleaned_name in ["Форма пром. атт.", "-"]:
            return None
            
        session = self.Session()
        try:
            format_control = session.execute(
                select(FormatControl).where(FormatControl.format_name == cleaned_name)
            ).scalar_one_or_none()
            
            if not format_control:
                format_control = FormatControl(format_name=cleaned_name)
                session.add(format_control)
                session.commit()
                logger.info(f"Создан новый формат контроля: {cleaned_name}")
            
            result = {
                'id': format_control.id,
                'name': format_control.format_name
            }
            return result
        finally:
            session.close()
    
    def _get_or_create_format_retests(self, format_name: str) -> Optional[Dict]:
        """Получить или создать формат переаттестации"""
        if not format_name:
            return None
            
        cleaned_name = self._clean_text(format_name)
        if not cleaned_name or cleaned_name in ["Форма пром. атт.", "-"]:
            return None
            
        session = self.Session()
        try:
            format_retests = session.execute(
                select(FormatRetests).where(FormatRetests.format_name == cleaned_name)
            ).scalar_one_or_none()
            
            if not format_retests:
                format_retests = FormatRetests(format_name=cleaned_name)
                session.add(format_retests)
                session.commit()
                logger.info(f"Создан новый формат переаттестации: {cleaned_name}")
            
            result = {
                'id': format_retests.id,
                'name': format_retests.format_name
            }
            return result
        finally:
            session.close()
    
    def parse_title_page(self) -> Tuple[Optional[Dict], Optional[Dict]]:
        """Парсинг титульной страницы"""
        if "Титул" not in self.wb.sheetnames:
            logger.error("Страница 'Титул' не найдена")
            return None, None
            
        sheet = self.wb["Титул"]
        
        # Ищем ячейку с "Направление подготовки"
        specialization_data = None
        for row in sheet.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and "Направление подготовки" in str(cell.value):
                    specialization_data = str(cell.value)
                    break
            if specialization_data:
                break
        
        if not specialization_data:
            logger.error("Не найдены данные о направлении подготовки")
            return None, None
        
        # Очищаем данные
        specialization_data = self._clean_text(specialization_data)
        logger.info(f"Найдены данные: {specialization_data}")
        
        # Парсим код и название специальности
        code_pattern = r'(\d{2}\.\d{2}\.\d{2})\s+([А-Яа-яA-Za-z\s\-]+?)(?:\s+направленность|\s*$)'
        code_match = re.search(code_pattern, specialization_data, re.IGNORECASE)
        
        specialization_code = None
        specialization_name = None
        
        if code_match:
            specialization_code = code_match.group(1)
            specialization_name = code_match.group(2).strip()
        else:
            alt_pattern = r'(\d{2}\.\d{2}\.\d{2})\s+([^,\.]+)'
            alt_match = re.search(alt_pattern, specialization_data)
            if alt_match:
                specialization_code = alt_match.group(1)
                specialization_name = alt_match.group(2).strip()
        
        # Парсим направленность (профиль)
        direction_pattern = r'направленность\s*\(профиль\)\s*программы\s*[:"\s]*([^"\)]+)'
        direction_match = re.search(direction_pattern, specialization_data, re.IGNORECASE)
        
        direction_name = None
        if direction_match:
            direction_name = direction_match.group(1).strip()
            direction_name = direction_name.strip('"').strip("'")
        
        if not specialization_code or not specialization_name:
            logger.error("Не удалось извлечь код или название специальности")
            return None, None
        
        # Получаем или создаем направление
        direction = None
        direction_id = None
        if direction_name:
            direction = self._get_or_create_direction(direction_name)
            direction_id = direction['id'] if direction else None
        
        # Получаем или создаем специализацию
        specialization = self._get_or_create_specialization(
            specialization_name, 
            specialization_code,
            direction_id
        )
        
        return specialization, direction
    
    def parse_pereattestatsiya_references(self):
        """Парсинг страницы Переаттестация для заполнения справочников"""
        if "Переаттестация" not in self.wb.sheetnames:
            logger.warning("Страница 'Переаттестация' не найдена")
            return [], [], []
        
        sheet = self.wb["Переаттестация"]
        
        study_programs = []
        format_controls = []
        format_retests = []
        
        # Поиск заголовков для определения столбцов
        headers = self._find_headers(sheet, {
            'name': ['Наименование', 'Дисциплина', 'Предмет'],
            'fact': ['Зачет результатов обучения', 'Факт'],
            'format': ['Форма пром. атт.', 'Форма контроля', 'Переаттестация']
        })
        
        logger.info(f"Найдены заголовки на странице Переаттестация: {headers}")
        
        # Парсим данные
        for row in sheet.iter_rows(min_row=6):
            name_cell = row[headers.get('name', 1) - 1] if headers.get('name') else row[1]
            fact_cell = row[headers.get('fact', 5) - 1] if headers.get('fact') else row[5]
            format_cell = row[headers.get('format', 8) - 1] if headers.get('format') else row[8]
            
            name = name_cell.value if name_cell.value else None
            fact = fact_cell.value if fact_cell.value else None
            format_name = format_cell.value if format_cell.value else None
            
            if not name or str(name) in ["Наименование", "Индекс", "-"]:
                continue
            
            # Сохраняем предмет
            study_program = self._get_or_create_study_program(str(name))
            if study_program:
                study_programs.append(study_program)
            
            # Сохраняем формат контроля (факт)
            if fact and str(fact) not in ["-", "None", ""]:
                format_control = self._get_or_create_format_control(str(fact))
                if format_control:
                    format_controls.append(format_control)
            
            # Сохраняем формат переаттестации
            if format_name and str(format_name) not in ["-", "None", ""]:
                format_retest = self._get_or_create_format_retests(str(format_name))
                if format_retest:
                    format_retests.append(format_retest)
        
        # Удаляем дубликаты по ID
        study_programs = list({p['id']: p for p in study_programs}.values())
        format_controls = list({f['id']: f for f in format_controls}.values())
        format_retests = list({r['id']: r for r in format_retests}.values())
        
        logger.info(f"Загружено предметов: {len(study_programs)}")
        logger.info(f"Загружено форматов контроля: {len(format_controls)}")
        logger.info(f"Загружено форматов переаттестации: {len(format_retests)}")
        
        return study_programs, format_controls, format_retests
    
    def _find_headers(self, sheet, header_patterns):
        """Поиск заголовков в таблице"""
        headers = {}
        
        for row in sheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip().lower()
                    
                    for key, patterns in header_patterns.items():
                        if key in headers:
                            continue
                        
                        for pattern in patterns:
                            if pattern.lower() in value:
                                headers[key] = cell.column
                                break
        
        return headers
    
    def parse_plan_page(self, specialization: Dict):
        """Парсинг страницы План для заполнения ControlTable"""
        if "План" not in self.wb.sheetnames:
            logger.warning("Страница 'План' не найдена")
            return []
        
        sheet = self.wb["План"]
        
        # Поиск заголовков для определения столбцов
        headers = self._find_headers(sheet, {
            'name': ['Наименование', 'Дисциплина', 'Предмет'],
            'hours_normal': ['По плану', 'Норма'],
            'hours_fact': ['Изучено и зачтено', 'Факт'],
            'format': ['Формы пром. атт.', 'Экзамен', 'Зачет']
        })
        
        logger.info(f"Найдены заголовки на странице План: {headers}")
        
        control_tables = []
        
        # Парсим данные
        for row in sheet.iter_rows(min_row=6):
            name_cell = row[headers.get('name', 1) - 1] if headers.get('name') else row[1]
            hours_normal_cell = row[headers.get('hours_normal', 4) - 1] if headers.get('hours_normal') else row[4]
            hours_fact_cell = row[headers.get('hours_fact', 7) - 1] if headers.get('hours_fact') else row[7]
            format_cell = row[headers.get('format', 3) - 1] if headers.get('format') else row[3]
            
            name = name_cell.value if name_cell.value else None
            hours_normal = hours_normal_cell.value if hours_normal_cell.value else None
            hours_fact = hours_fact_cell.value if hours_fact_cell.value else None
            format_name = format_cell.value if format_cell.value else None
            
            if not name or str(name) in ["Наименование", "Индекс", "-"]:
                continue
            
            # Получаем или создаем предмет
            study_program = self._get_or_create_study_program(str(name))
            if not study_program:
                continue
            
            # Получаем или создаем формат контроля (норма)
            format_control_norma = None
            if format_name and str(format_name) not in ["-", "None", ""]:
                format_clean = str(format_name).split()[0] if format_name else None
                if format_clean:
                    format_control_norma = self._get_or_create_format_control(format_clean)
            
            # Создаем запись в ControlTable
            session = self.Session()
            try:
                # Проверяем, нет ли уже такой записи
                existing = session.execute(
                    select(ControlTable).where(
                        ControlTable.specialization_id == specialization['id'],
                        ControlTable.study_program_id == study_program['id']
                    )
                ).scalar_one_or_none()
                
                if not existing:
                    control_table = ControlTable(
                        specialization_id=specialization['id'],
                        study_program_id=study_program['id'],
                        format_control_norma_id=format_control_norma['id'] if format_control_norma else None,
                        hours_normal=str(hours_normal) if hours_normal else None,
                        hours_fact=str(hours_fact) if hours_fact else None
                    )
                    session.add(control_table)
                    session.commit()
                    control_tables.append({
                        'id': control_table.id,
                        'specialization_id': control_table.specialization_id,
                        'study_program_id': control_table.study_program_id
                    })
                    logger.info(f"Создана запись ControlTable для предмета: {name}")
                else:
                    # Обновляем существующую запись
                    updated = False
                    if hours_normal and not existing.hours_normal:
                        existing.hours_normal = str(hours_normal)
                        updated = True
                    if hours_fact and not existing.hours_fact:
                        existing.hours_fact = str(hours_fact)
                        updated = True
                    
                    if updated:
                        session.commit()
                        logger.info(f"Обновлена запись ControlTable для предмета: {name}")
            finally:
                session.close()
        
        logger.info(f"Создано/обновлено записей ControlTable из плана: {len(control_tables)}")
        return control_tables
    
    def parse_retests_page_for_control(self, specialization: Dict):
        """Парсинг страницы Переаттестация для заполнения ControlTable"""
        if "Переаттестация" not in self.wb.sheetnames:
            return []
        
        sheet = self.wb["Переаттестация"]
        
        # Поиск заголовков
        headers = self._find_headers(sheet, {
            'name': ['Наименование', 'Дисциплина', 'Предмет'],
            'hours_fact': ['з.е.', 'Часов', 'Факт'],
            'format_retests': ['Форма пром. атт.', 'Переаттестация']
        })
        
        logger.info(f"Найдены заголовки на странице Переаттестация: {headers}")
        
        updated_count = 0
        
        # Парсим данные
        session = self.Session()
        try:
            for row in sheet.iter_rows(min_row=6):
                name_cell = row[headers.get('name', 1) - 1] if headers.get('name') else row[1]
                hours_fact_cell = row[headers.get('hours_fact', 6) - 1] if headers.get('hours_fact') else row[6]
                format_retests_cell = row[headers.get('format_retests', 8) - 1] if headers.get('format_retests') else row[8]
                
                name = name_cell.value if name_cell.value else None
                hours_fact = hours_fact_cell.value if hours_fact_cell.value else None
                format_retests_name = format_retests_cell.value if format_retests_cell.value else None
                
                if not name or str(name) in ["Наименование", "Индекс", "-"]:
                    continue
                
                # Находим предмет
                cleaned_name = self._clean_text(str(name))
                study_program = session.execute(
                    select(StudyProgram).where(StudyProgram.name == cleaned_name)
                ).scalar_one_or_none()
                
                if not study_program:
                    continue
                
                # Находим формат переаттестации
                format_retests = None
                if format_retests_name and str(format_retests_name) not in ["-", "None", ""]:
                    cleaned_format = self._clean_text(str(format_retests_name))
                    format_retests = session.execute(
                        select(FormatRetests).where(FormatRetests.format_name == cleaned_format)
                    ).scalar_one_or_none()
                
                # Обновляем или создаем запись в ControlTable
                control_table = session.execute(
                    select(ControlTable).where(
                        ControlTable.specialization_id == specialization['id'],
                        ControlTable.study_program_id == study_program.id
                    )
                ).scalar_one_or_none()
                
                if control_table:
                    # Обновляем существующую запись
                    updated = False
                    if hours_fact and not control_table.hours_fact:
                        control_table.hours_fact = str(hours_fact)
                        updated = True
                    if format_retests and not control_table.format_retests_id:
                        control_table.format_retests_id = format_retests.id
                        updated = True
                    
                    if updated:
                        updated_count += 1
                else:
                    # Создаем новую запись
                    control_table = ControlTable(
                        specialization_id=specialization['id'],
                        study_program_id=study_program.id,
                        hours_fact=str(hours_fact) if hours_fact else None,
                        format_retests_id=format_retests.id if format_retests else None
                    )
                    session.add(control_table)
                    updated_count += 1
            
            if updated_count > 0:
                session.commit()
        finally:
            session.close()
        
        logger.info(f"Обновлено/создано записей из переаттестации: {updated_count}")
    
    def parse_all(self) -> Dict[str, Any]:
        """Основной метод парсинга"""
        logger.info(f"Начинаем парсинг файла: {self.file_path}")
        
        result = {
            'specialization': None,
            'direction': None,
            'study_programs': 0,
            'format_controls': 0,
            'format_retests': 0,
            'control_tables': 0
        }
        
        # 1. Парсим титульную страницу
        specialization, direction = self.parse_title_page()
        if not specialization:
            logger.error("Не удалось получить специализацию, прерываем парсинг")
            return result
        
        result['specialization'] = specialization
        result['direction'] = direction
        
        # 2. Парсим страницу Переаттестация для справочников
        study_programs, format_controls, format_retests = self.parse_pereattestatsiya_references()
        result['study_programs'] = len(study_programs)
        result['format_controls'] = len(format_controls)
        result['format_retests'] = len(format_retests)
        
        # 3. Парсим страницу План для ControlTable
        control_tables = self.parse_plan_page(specialization)
        result['control_tables'] = len(control_tables)
        
        # 4. Парсим страницу Переаттестация для обновления ControlTable
        self.parse_retests_page_for_control(specialization)
        
        logger.info(f"Парсинг завершен успешно! Результаты: {result}")
        return result


def parse_xlsx_file(file_path: str, db_url: str) -> Dict[str, Any]:
    """Функция для вызова из API"""
    parser = XLSXParser(file_path, db_url)
    return parser.parse_all()